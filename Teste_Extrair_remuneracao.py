import pandas as pd
import holidays
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Janela oculta só para abrir diálogos
root = tk.Tk()
root.withdraw()

def formatar_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    # --- Cabeçalho ---
    header_fill = PatternFill("solid", fgColor="1F4E78")  # azul escuro
    header_font = Font(color="FFFFFF", bold=True)         # branco, negrito

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "@"   # tratar como texto para não dar erro no LibreOffice

    # --- Ajuste automático da largura ---
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 2

    ws.column_dimensions["A"].width = 19
    
    wb.save(caminho)


# --- 1) Selecionar arquivo CSV ---
"""INPUT = filedialog.askopenfilename(
    title="Selecione o arquivo CSV",
    filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
)

if not INPUT:
    print("Nenhum arquivo selecionado.")
    quit()"""

# --- 2) Escolher onde salvar ---
"""OUTPUT_EXCEL = filedialog.asksaveasfilename(
    title="Salvar arquivo Excel",
    defaultextension=".xlsx",
    filetypes=[("Excel", "*.xlsx")]
)

if not OUTPUT_EXCEL:
    print("Nenhum local de salvamento escolhido.")
    quit()"""
def cct_extrair(vlr):
    
    
    ########   USANDO CODIGO DA REMUNERACAO PARA FINS DE TESTE FUTURAMENTE TROCAR#############
    
    
    INPUT = filedialog.askopenfilename(
    title="Selecione o arquivo da CCT",
    filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
    )

    if not INPUT:
        print("Nenhum arquivo selecionado.")
        quit()
    # --- 3) PROCESSAMENTO DO ARQUIVO (seu código abaixo) ---

    # 1) Ler CSV
    df = pd.read_csv(INPUT)
    df.columns = df.columns.str.strip()

    # 2) Converter coluna de data
    df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')

    # 3) Feriados do Brasil
    anos = df['DATA'].dt.year.dropna().unique().astype(int)
    feriados = holidays.Brazil(state="PE",years=anos)
    feriados_set = set(feriados.keys())

    # 4) Filtrar dias úteis
    mask = (df['DATA'].dt.weekday < 5) & (~df['DATA'].dt.date.isin(feriados_set))
    df_uteis = df[mask].copy()

    # 5) Garantir que DIAPASST seja numérico
    df_uteis['DIAPASST'] = pd.to_numeric(df_uteis['DIAPASST'], errors='coerce')

    # 6) Agrupar por CDOPERADOR e CDLINHA
    resultado = df_uteis.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['DIAPASST'].mean()

    # 7) Renomear coluna da média
    resultado.rename(columns={'DIAPASST': 'MediaTotal'}, inplace=True)

    # 8) Arredondar média
    resultado['MediaTotal'] = resultado['MediaTotal'].round(2)

    if vlr == 1:
        return resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy()
    else:
        OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not OUTPUT_EXCEL:
            print("Nenhum local de salvamento escolhido.")
            quit()
        resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)

def remuneracao_Extrair(vlr):
    INPUT = filedialog.askopenfilename(
    title="Selecione o arquivo da REM",
    filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
    )

    if not INPUT:
        print("Nenhum arquivo selecionado.")
        quit()
    # --- 3) PROCESSAMENTO DO ARQUIVO (seu código abaixo) ---

    # 1) Ler CSV
    df = pd.read_csv(INPUT)
    df.columns = df.columns.str.strip()

    # 2) Converter coluna de data
    df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')

    # 3) Feriados do Brasil
    anos = df['DATA'].dt.year.dropna().unique().astype(int)
    feriados = holidays.Brazil(state="PE",years=anos)
    feriados_set = set(feriados.keys())

    # 4) Filtrar dias úteis
    mask = (df['DATA'].dt.weekday < 5) & (~df['DATA'].dt.date.isin(feriados_set))
    df_uteis = df[mask].copy()

    # 5) Garantir que DIAPASST seja numérico
    df_uteis['DIAPASST'] = pd.to_numeric(df_uteis['DIAPASST'], errors='coerce')

    # 6) Agrupar por CDOPERADOR e CDLINHA
    resultado = df_uteis.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['DIAPASST'].mean()

    # 7) Renomear coluna da média
    resultado.rename(columns={'DIAPASST': 'MediaTotal'}, inplace=True)

    # 8) Arredondar média
    resultado['MediaTotal'] = resultado['MediaTotal'].round(2)

    if vlr == 1:
        return resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy()
    else:
        OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not OUTPUT_EXCEL:
            print("Nenhum local de salvamento escolhido.")
            quit()
        resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)



def ambos_cct_rem():
    dados_cct = cct_extrair(1)
    dados_remune = remuneracao_Extrair(1)

    resultado_combinado = pd.concat([dados_cct, dados_remune], ignore_index=True)
    resultado_ordenado = resultado_combinado.sort_values(
    by=["CDOPERADOR", "CDLINHA"]
    )
    OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
    )
    
    if not OUTPUT_EXCEL:
        print("Nenhum local de salvamento escolhido.")
        quit()
        
    resultado_ordenado.to_excel(OUTPUT_EXCEL, index=False)


#j = remuneracao_Extrair(INPUT, vlr=0)
#print(j)

def menu():
    mmenu = input("""
╔════════════════════════════════════════╗
║           SELECIONE UMA OPÇÃO          ║
╠════════════════════════════════════════╣
║ [1]  Arquivo CTT                       ║
║ [2]  Arquivo Remuneração               ║
║ [3]  Ambos (CTT + Remuneração)         ║
╚════════════════════════════════════════╝
Digite a opção desejada: """)
    if mmenu == '1':
        cct_extrair(0)
    elif mmenu == '2':
        remuneracao_Extrair(0)
    elif mmenu == '3':
        ambos_cct_rem()
    else:
        print('\nOPÇÃO INVALIDA')
        menu()
menu()