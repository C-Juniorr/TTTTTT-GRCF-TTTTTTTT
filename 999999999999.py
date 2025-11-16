import os
import pandas as pd
import holidays
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# para gerar PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet

# Janela oculta só para abrir diálogos
root = tk.Tk()
root.withdraw()

def formatar_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    # --- Cabeçalho ---
    header_fill = PatternFill("solid", fgColor="1F4E78")  # azul escuro
    header_font = Font(color="FFFFFF", bold=True)         # branco, negrito

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "@"   # tratar como texto para não dar erro no LibreOffice

    # --- Ajuste automático da largura ---
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2

    ws.column_dimensions["A"].width = 19
    
    wb.save(caminho)

def save_df_as_pdf(df: pd.DataFrame, pdf_path: str, title: str = None):
    """
    Salva um DataFrame em um PDF simples usando reportlab.
    """
    # garante que exista pelo menos uma linha/coluna
    if df.shape[0] == 0 or df.shape[1] == 0:
        # cria um PDF vazio com mensagem
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Sem dados para exportar.", styles["Normal"]))
        doc.build(story)
        return

    # prepara dados para a tabela (primeira linha = cabeçalho)
    data = [list(df.columns)]
    # converter valores None para string vazia e garantir conversão para string
    for row in df.itertuples(index=False):
        data.append([("" if v is None else str(v)) for v in row])

    # usar paisagem se muitas colunas
    pagesize = landscape(A4) if df.shape[1] > 5 else A4

    doc = SimpleDocTemplate(pdf_path, pagesize=pagesize, leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    if title:
        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 12))

    # criar tabela
    table = Table(data, repeatRows=1)  # repeatRows=1 repete cabeçalho em páginas seguintes
    # estilo simples
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),  # cabeçalho azul
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ])
    table.setStyle(style)

    story.append(table)
    doc.build(story)

def inputt(titl):
    INPUT = filedialog.askopenfilename(
        title=titl,
        filetypes=[("Arquivos CSV", "*.csv"),("Arquivos TXT", "*.txt"), ("Todos os arquivos", "*.*")]
    )

    if not INPUT:
        print("Nenhum arquivo selecionado.")
        quit()
        
    return INPUT

def output():
    OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

    if not OUTPUT_EXCEL:
        print("Nenhum local de salvamento escolhido.")
        quit()
    return OUTPUT_EXCEL

def cct_extrair(vlr):
    INPUT = inputt("Selecione o arquivo da CCT")
    # 1) Ler CSV
    df = pd.read_csv(INPUT)
    df.columns = df.columns.str.strip()

    # 2) Converter coluna de data
    df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')

    # 3) Feriados do Brasil
    anos = df['DATA'].dt.year.dropna().unique().astype(int)
    feriados = holidays.Brazil(state="PE", years=anos)
    feriados_set = set(feriados.keys())

    # 4) Filtrar dias úteis
    mask = (df['DATA'].dt.weekday < 5) & (~df['DATA'].dt.date.isin(feriados_set))
    df_uteis = df[mask].copy()
    
    # 5) Garantir que DIAPASST seja numérico
    df_uteis['DIAPASST'] = pd.to_numeric(df_uteis['DIAPASST'], errors='coerce')

    # 6) Agrupar por CDOPERADOR e CDLINHA
    resultado = df_uteis.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['DIAPASST'].mean()

    # 7) Renomear coluna da média
    resultado.rename(columns={'DIAPASST': 'MediaTotal'}, inplace=True)

    # 8) Arredondar média
    resultado['MediaTotal'] = resultado['MediaTotal'].round(2)

    if vlr == 1:
        return resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy()
    else:
        OUTPUT_EXCEL = output()
        resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)

        # salvar PDF com mesmo nome (mesmo diretório)
        pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
        titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
        save_df_as_pdf(resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
        print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
        print(f"Arquivo PDF salvo em: {pdf_path}")

def remuneracao_Extrair(vlr):
    INPUT = inputt("Selecione o arquivo da Remuneracao")
    # 1) Ler CSV
    df = pd.read_csv(INPUT)
    df.columns = df.columns.str.strip()

    # 2) Converter coluna de data
    df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')

    # 3) Feriados do Brasil
    anos = df['DATA'].dt.year.dropna().unique().astype(int)
    feriados = holidays.Brazil(state="PE", years=anos)
    feriados_set = set(feriados.keys())

    # 4) Filtrar dias úteis
    mask = (df['DATA'].dt.weekday < 5) & (~df['DATA'].dt.date.isin(feriados_set))
    df_uteis = df[mask].copy()

    # 5) Garantir que DIAPASST seja numérico
    df_uteis['DIAPASST'] = pd.to_numeric(df_uteis['DIAPASST'], errors='coerce')

    # 6) Agrupar por CDOPERADOR e CDLINHA
    resultado = df_uteis.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['DIAPASST'].mean()

    # 7) Renomear coluna da média
    resultado.rename(columns={'DIAPASST': 'MediaTotal'}, inplace=True)

    # 8) Arredondar média
    resultado['MediaTotal'] = resultado['MediaTotal'].round(2)

    if vlr == 1:
        return resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy()
    else:
        OUTPUT_EXCEL = output()
        resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)

        # salvar PDF com mesmo nome (mesmo diretório)
        pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
        titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
        save_df_as_pdf(resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
        print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
        print(f"Arquivo PDF salvo em: {pdf_path}")

def ambos_cct_rem():
    dados_cct = cct_extrair(1)
    dados_remune = remuneracao_Extrair(1)

    resultado_combinado = pd.concat([dados_cct, dados_remune], ignore_index=True)
    resultado_ordenado = resultado_combinado.sort_values(
        by=["CDOPERADOR", "CDLINHA"]
    )
    OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
    )
    
    if not OUTPUT_EXCEL:
        print("Nenhum local de salvamento escolhido.")
        quit()
        
    resultado_ordenado.to_excel(OUTPUT_EXCEL, index=False)
    formatar_excel(OUTPUT_EXCEL)

    # salvar PDF também
    pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
    titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
    save_df_as_pdf(resultado_ordenado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
    print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
    print(f"Arquivo PDF salvo em: {pdf_path}")

def menu():
    mmenu = input("""
╔════════════════════════════════════════╗
║           SELECIONE UMA OPÇÃO          ║
╠════════════════════════════════════════╣
║ [1]  Arquivo CTT                       ║
║ [2]  Arquivo Remuneração               ║
║ [3]  Ambos (CTT + Remuneração)         ║
╚════════════════════════════════════════╝
Digite a opção desejada: """)
    if mmenu == '1':
        cct_extrair(0)
    elif mmenu == '2':
        remuneracao_Extrair(0)
    elif mmenu == '3':
        ambos_cct_rem()
    else:
        print('\nOPÇÃO INVALIDA')
        menu()

if __name__ == "__main__":
    menu()
