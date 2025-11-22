import os
import pandas as pd
import tkinter as tk
from tkinter import Tk,filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# para gerar PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet

#EXTRAIR 1 2 e ambos OK
#FILTRO OK


#CRIAR UMA PLANILHA PARA PEGAR O DIA A DIA


#atualizar codigo para pegar por linha


root = tk.Tk()
root.withdraw()

def formatar_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    # --- Cabeçalho ---
    header_fill = PatternFill("solid", fgColor="1F4E78")  # azul escuro
    header_font = Font(color="FFFFFF", bold=True)         # branco, negrito

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "@"   # tratar como texto para não dar erro no LibreOffice
    ws.row_dimensions[1].height = 25 # Ajuste este valor conforme sua necessidade
    
    # --- Ajuste automático da largura ---
    
    for col in ws.columns:
        letra_coluna = col[0].column_letter
        tamanho_titulo = len(str(ws[letra_coluna + '1'].value)) if ws[letra_coluna + '1'].value is not None else 0
        tamanho_max_conteudo = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[letra_coluna].width = max(tamanho_titulo, tamanho_max_conteudo) + 2
    # --- Ajuste automático da largura ---
    #for col in ws.columns:
    #    ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
    for cell in ws['D']:   # Coluna onde está TTdia
        cell.alignment = Alignment(horizontal='right')

    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 15
    wb.save(caminho)

def save_df_as_pdf(df: pd.DataFrame, pdf_path: str, title: str = None):
    """
    Salva um DataFrame em um PDF simples usando reportlab.
    """
    # garante que exista pelo menos uma linha/coluna
    if df.shape[0] == 0 or df.shape[1] == 0:
        # cria um PDF vazio com mensagem
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Sem dados para exportar.", styles["Normal"]))
        doc.build(story)
        return

    # prepara dados para a tabela (primeira linha = cabeçalho)
    data = [list(df.columns)]
    # converter valores None para string vazia e garantir conversão para string
    for row in df.itertuples(index=False):
        data.append([("" if v is None else str(v)) for v in row])

    # usar paisagem se muitas colunas
    pagesize = landscape(A4) if df.shape[1] > 5 else A4

    doc = SimpleDocTemplate(pdf_path, pagesize=pagesize, leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    if title:
        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 12))

    # criar tabela
    table = Table(data, repeatRows=1)  # repeatRows=1 repete cabeçalho em páginas seguintes
    # estilo simples
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),  # cabeçalho azul
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ])
    table.setStyle(style)

    story.append(table)
    doc.build(story)

def inputt(titl):

    INPUT = filedialog.askopenfilename(
        title=titl,
        filetypes=[("Arquivos TXT", "*.txt"),("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
    )

    if not INPUT:
        print("Nenhum arquivo selecionado.")
        quit()
        
    return INPUT

def output():
    OUTPUT_EXCEL = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

    if not OUTPUT_EXCEL:
        print("Nenhum local de salvamento escolhido.")
        quit()
    return OUTPUT_EXCEL

def cct_extrair(vlr):
    lista_operadores = input("Digite os operadores separados por vírgula exemplo BOA, CAX, CSR, EME, GLO, SJT, VML, DE ENTER SE QUISER TODOS: ")
    


    lista_linhas = input("Digite as linhas separadas por , exemplo: 101, 200, 400,: ")
    


    INPUT = inputt("Selecione o arquivo da CCT")
    colunas_que_quero = ["CDOPERADOR", "CDLINHA", "DATA", "DIAPASST"]

    df = pd.read_csv(INPUT)
    #df = pd.read_csv(INPUT, sep='\t', engine='python')
    #df = pd.read_csv(INPUT, sep=";", engine="python", encoding="latin-1")

    df = df[colunas_que_quero]
    # FILTRO LINHAS
    if lista_linhas.strip():
        
        lista_linhas = [ln.strip() for ln in lista_linhas.split(",") if ln.strip()]
        df["CDLINHA"] = df["CDLINHA"].astype(str).str.strip()
        df = df[df["CDLINHA"].isin(lista_linhas)]
    
    elif lista_operadores.strip():
        lista_operadores = [op.strip() for op in lista_operadores.split(",") if op.strip()]
        df = df[df["CDOPERADOR"].isin(lista_operadores)]


        df["CDLINHA"] = df["CDLINHA"].astype(str).str.strip()
        df = df[df["CDLINHA"].isin(lista_linhas)]

    df["NMEFETPASST"] = df["NMEFETPASST"].str.replace(',', '.').str.strip()
    df["NMEFETPASST"] = pd.to_numeric(df["NMEFETPASST"], errors='coerce')

    df_dia_a_dia = df[["CDOPERADOR", "CDLINHA", "DTOPERACAO", "NMEFETPASST"]].copy()
    df_dia_a_dia = df_dia_a_dia.sort_values(by=["CDOPERADOR", "CDLINHA", "DTOPERACAO"])

    df_dia_a_dia['NMEFETPASST'] = df_dia_a_dia['NMEFETPASST'].round(0).astype(int)

    df = df[df["DSDIATIPO"] == "DUT"]


    #df["NMEFETPASST"] = df["NMEFETPASST"].str.replace(',', '.').str.strip()
    #df["NMEFETPASST"] = pd.to_numeric(df["NMEFETPASST"], errors='coerce')
    
    #df_dia_a_dia = df[["CDOPERADOR", "CDLINHA", "DTOPERACAO", "NMEFETPASST"]].copy()
    #df_dia_a_dia = df_dia_a_dia.sort_values(by=["CDOPERADOR", "CDLINHA", "DTOPERACAO"])

    # SALVAR A PLANILHA DIA A DIA

    df_soma = df.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['NMEFETPASST'].sum()


    # Renomear a coluna para deixar mais claro
    df_soma = df_soma.rename(columns={'NMEFETPASST': 'TOTAL_NMEFETPASST'})


    df_contagem = df.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['NMEFETPASST'].count()
    df_contagem = df_contagem.rename(columns={'NMEFETPASST': 'QTD_OCORRENCIAS'})

    # 2. Juntar com df_soma que já tem a soma total
    df_media = pd.merge(df_soma, df_contagem, on=['CDOPERADOR', 'CDLINHA'])

    # 3. Calcular a média dividindo o total pela quantidade
    df_media['MediaTotal'] = df_media['TOTAL_NMEFETPASST'] / df_media['QTD_OCORRENCIAS']


    df_para_salvar = df_media[['CDOPERADOR', 'CDLINHA', 'MediaTotal']]

    # Ordenar por CDOPERADOR e, se quiser, também por CDLINHA
    df_para_salvar = df_para_salvar.sort_values(by=['CDOPERADOR', 'CDLINHA'])

    # Tirar as casas decimais da média (arredondando para inteiro)
    df_para_salvar['MediaTotal'] = df_para_salvar['MediaTotal'].round(0).astype(int)

    if vlr == 1:
        if lista_linhas.strip():
            return df_para_salvar[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy(), df_dia_a_dia.copy(), INPUT, lista_linhas
        elif lista_operadores.strip():
            return df_para_salvar[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy(), df_dia_a_dia.copy(), INPUT, lista_operadores
        else:
            return df_para_salvar[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy(), df_dia_a_dia.copy(), INPUT
    else:
        OUTPUT_EXCEL = output()
        df_para_salvar[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)

        saida_dia_a_dia = os.path.splitext(OUTPUT_EXCEL)[0] + "_DIA_A_DIA.xlsx"
        df_dia_a_dia.to_excel(saida_dia_a_dia, index=False)
        formatar_excel(saida_dia_a_dia)


        # salvar PDF com mesmo nome (mesmo diretório)
        '''pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
        titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
        save_df_as_pdf(df_para_salvar[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
        print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
        print(f"Arquivo PDF salvo em: {pdf_path}")'''


def remuneracao_Extrair(vlr, caminho, listafiltro=[]):

    print(listafiltro)
    if listafiltro == []:
        print("Nenhum filtro")
        lista_linhas = input("Digite as linhas separadas por , exemplo: 101, 200, 400,: ")
        lista_operadores = input("Digite os operadores separados por vírgula exemplo BOA, CAX, CSR, EME, GLO, SJT, VML, DE ENTER SE QUISER TODOS: ")
        
    elif all(item.isdigit() for item in listafiltro):
        print("Filtro é de LINHAS")
        lista_operadores = listafiltro
    else:
        print("Filtro é de OPERADORES")
        lista_linhas = listafiltro
        
    if vlr == 0:
        arquivo_dut = inputt("Selecione o arquivo que contém as datas DUT")
        df_dut = pd.read_csv(arquivo_dut, sep='\t', engine='python')
        df_dut["DTOPERACAO"] = pd.to_datetime(df_dut["DTOPERACAO"], format="%d/%m/%Y", errors='coerce')
        df_dut = df_dut[df_dut["DSDIATIPO"] == "DUT"].dropna(subset=["DTOPERACAO"])
        lista_datas_dut = sorted(df_dut["DTOPERACAO"].dt.date.unique())
    else:
        df_dut = pd.read_csv(caminho, sep='\t', engine='python')
        df_dut["DTOPERACAO"] = pd.to_datetime(df_dut["DTOPERACAO"], format="%d/%m/%Y", errors='coerce')
        df_dut = df_dut[df_dut["DSDIATIPO"] == "DUT"].dropna(subset=["DTOPERACAO"])
        lista_datas_dut = sorted(df_dut["DTOPERACAO"].dt.date.unique())


    INPUT = inputt("Selecione o arquivo da Remuneracao")
    df = pd.read_csv(INPUT, sep='\t', engine='python')
    
    if lista_linhas.strip():   
        lista_linhas = [ln.strip() for ln in lista_linhas.split(",") if ln.strip()]
        df["CDLINHA"] = df["CDLINHA"].astype(str).str.strip()
        df = df[df["CDLINHA"].isin(lista_linhas)]
    
    elif lista_operadores.strip():
        lista_operadores = [op.strip() for op in lista_operadores.split(",") if op.strip()]
        df = df[df["CDOPERADOR"].isin(lista_operadores)]
        

    df.columns = df.columns.str.strip()

    # Aqui, substituir 'DATA' por 'DTOPERACAO' se for o caso
    df['DTOPERACAO'] = pd.to_datetime(df['DTOPERACAO'], dayfirst=True, errors='coerce')
    
    #df['DTOPERACAO'] = pd.to_datetime(df['DTOPERACAO'], dayfirst=True, errors='coerce').dt.date
   
    df_uteis = df[df['DTOPERACAO'].dt.date.isin(lista_datas_dut)].copy()
    
    df_uteis['NMPASSTOTAL'] = (
    df_uteis['NMPASSTOTAL']
    .astype(str)
    .str.replace(',', '.', regex=False)
    .str.strip()
    )
    df_uteis['DIAPASST'] = pd.to_numeric(df_uteis['NMPASSTOTAL'], errors='coerce')

    df_dia_a_dia = df[["CDOPERADOR", "CDLINHA", "DTOPERACAO", "NMPASSTOTAL"]].copy()
    df_dia_a_dia = df_dia_a_dia.sort_values(by=["CDOPERADOR", "CDLINHA", "DTOPERACAO"])

    resultado = df_uteis.groupby(['CDOPERADOR', 'CDLINHA'], as_index=False)['DIAPASST'].mean()
    resultado.rename(columns={'DIAPASST': 'MediaTotal'}, inplace=True)
    resultado['MediaTotal'] = resultado['MediaTotal'].round(0).astype(int)

    if vlr == 1 or vlr == 2:
        return resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].copy(), df_dia_a_dia.copy()
    else:
        OUTPUT_EXCEL = output()
        resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']].to_excel(OUTPUT_EXCEL, index=False)
        formatar_excel(OUTPUT_EXCEL)
        
        saida_dia_a_dia = os.path.splitext(OUTPUT_EXCEL)[0] + "_DIA_A_DIA.xlsx"
        df_dia_a_dia['DTOPERACAO'] = pd.to_datetime(df_dia_a_dia['DTOPERACAO'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
        df_dia_a_dia.to_excel(saida_dia_a_dia, index=False)
        formatar_excel(saida_dia_a_dia)

        '''pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
        titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
        save_df_as_pdf(resultado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
        print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
        print(f"Arquivo PDF salvo em: {pdf_path}")'''

def ambos_cct_rem():
    
    dados_cct = cct_extrair(1)
    lista_filtro = dados_cct[3]
    dados_remune = remuneracao_Extrair(1, dados_cct[2],lista_filtro)

    resultado_combinado = pd.concat([dados_cct[0], dados_remune[0]], ignore_index=True)
    resultado_ordenado = resultado_combinado.sort_values(
        by=["CDOPERADOR", "CDLINHA"]
    )

    r1 = dados_cct[1].rename(columns={'NMEFETPASST': 'TTDIA'})
    r2 = dados_remune[1].rename(columns={'NMPASSTOTAL': 'TTDIA'})
    r2['DTOPERACAO'] = pd.to_datetime(r2['DTOPERACAO'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
    #r2['DTOPERACAO'] = pd.to_datetime(r2['DTOPERACAO'], dayfirst=True).dt.strftime('%d/%m/%Y')

    #r2['DTOPERACAO'] = r2['DTOPERACAO'].dt.date

    resultado_dia_dia = pd.concat([r1, r2], ignore_index=True)
    #resultado_dia_dia["TTDIA"] = resultado_dia_dia["TTDIA"].round(0).astype(int)
    resultado_dia_dia = resultado_dia_dia.sort_values(by=["CDOPERADOR", "CDLINHA", "DTOPERACAO"])

    OUTPUT_EXCEL = output()
        
    resultado_ordenado.to_excel(OUTPUT_EXCEL, index=False)
    formatar_excel(OUTPUT_EXCEL)

    OUTPUT_DIA_DIA = os.path.splitext(OUTPUT_EXCEL)[0] + "_DIA_A_DIA.xlsx"
    resultado_dia_dia.to_excel(OUTPUT_DIA_DIA, index=False)
    formatar_excel(OUTPUT_DIA_DIA)

    # salvar PDF também
    pdf_path = os.path.splitext(OUTPUT_EXCEL)[0] + ".pdf"
    titulo_pdf = os.path.splitext(os.path.basename(OUTPUT_EXCEL))[0]
    #save_df_as_pdf(resultado_ordenado[['CDOPERADOR', 'CDLINHA', 'MediaTotal']], pdf_path, title=titulo_pdf)
    save_df_as_pdf(resultado_dia_dia[['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'TTDIA']], pdf_path, title=titulo_pdf)
    print(f"Arquivo Excel salvo em: {OUTPUT_EXCEL}")
    print(f"Arquivo PDF salvo em: {pdf_path}")



def menu():
    mmenu = input("""
╔════════════════════════════════════════╗
║           SELECIONE UMA OPÇÃO          ║
╠════════════════════════════════════════╣
║ [1]  Arquivo CTT                       ║
║ [2]  Arquivo Remuneração               ║
║ [3]  Ambos (CTT + Remuneração)         ║
║ [0]  Sair                              ║
╚════════════════════════════════════════╝
Digite a opção desejada: """)
    if mmenu == '1':
        cct_extrair(0)
    elif mmenu == '2':
        remuneracao_Extrair(0, '')
    elif mmenu == '3':
        ambos_cct_rem()
    elif mmenu == '0':
        quit()
    else:
        print('\nOPÇÃO INVALIDA')
        menu()

if __name__ == "__main__":
    menu()
